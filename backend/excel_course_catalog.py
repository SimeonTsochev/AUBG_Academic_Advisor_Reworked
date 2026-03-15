from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
import logging
import re

courses_by_code: Dict[str, Dict[str, Any]] = {}

_courses_sorted: List[Dict[str, Any]] = []
_loaded_path: Optional[str] = None
_loaded_mtime: Optional[float] = None
_catalog_payload: Dict[str, Any] = {
    "courses": [],
    "by_code": {},
    "codes": [],
}

_CODE_RE = re.compile(r"([A-Z]{2,4})\s*[-/]?\s*(\d{3,4}[A-Z]?)")
_CREDITS_RE = re.compile(r"credits?\s*:\s*([0-9]+(?:\.[0-9]+)?)\s*CR", re.IGNORECASE)
_TAG_SPLIT_RE = re.compile(r"[;\r\n]+")
_SPACE_RE = re.compile(r"\s+")

_HEADER_ALIASES = {
    "department": {"department", "dept", "prefix", "subject"},
    "course": {"course", "course number", "number"},
    "code": {"code", "course code"},
    "title": {"label", "title", "course title", "name", "course name"},
    "credits": {"credits", "credit"},
    "level": {"level", "course level"},
    "area_of_study": {"area of study", "area"},
    "course_notes": {"course notes", "notes", "description"},
}

_TERM_HEADER_HINTS = ("term", "semester", "availability", "offered")

_GEN_ED_STANDALONE_TAGS = {
    "aesthetic expression",
    "historical sources",
    "historical research",
    "moral and philosophical reasoning",
    "quantitative reasoning",
    "scientific investigation",
    "principles of textual analysis",
    "case studies in textual analysis",
}


def _normalize_space(text: str) -> str:
    return _SPACE_RE.sub(" ", text).strip()


def _clean_header(value: object) -> str:
    return _normalize_space(str(value or ""))


def _normalize_header_key(value: object) -> str:
    text = _clean_header(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return _normalize_space(str(value))


def _raw_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value)


def normalize_course_code(code: str) -> str:
    text = _clean_cell(code).upper()
    if not text:
        return ""
    match = _CODE_RE.search(text.replace(".", ""))
    if not match:
        return text
    return f"{match.group(1)} {match.group(2)}"


def _code_from_parts(department: str, course: str, code: str) -> str:
    normalized_code = normalize_course_code(code)
    if normalized_code:
        return normalized_code
    department_clean = re.sub(r"[^A-Z]", "", _clean_cell(department).upper())
    course_clean = re.sub(r"[^0-9A-Z]", "", _clean_cell(course).upper())
    if not department_clean or not course_clean:
        return ""
    return normalize_course_code(f"{department_clean} {course_clean}")


def _split_tags(value: object) -> List[str]:
    raw = _raw_cell(value)
    if not raw.strip():
        return []
    parts = _TAG_SPLIT_RE.split(raw)
    tags: List[str] = []
    for part in parts:
        cleaned = _normalize_space(part)
        if cleaned:
            tags.append(cleaned)
    return list(dict.fromkeys(tags))


def _parse_numeric(value: object) -> Optional[float]:
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_cell(value)
    if not text:
        return None
    match = re.search(r"[0-9]+(?:\.[0-9]+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _credits_from_value_or_notes(credits_value: object, notes_value: object) -> Optional[int]:
    parsed = _parse_numeric(credits_value)
    if parsed is None:
        notes = _raw_cell(notes_value)
        match = _CREDITS_RE.search(notes)
        if match:
            parsed = float(match.group(1))
    if parsed is None or parsed <= 0:
        return None
    if float(parsed).is_integer():
        return int(parsed)
    return int(round(parsed))


def _gen_ed_from_area_tags(area_tags: List[str]) -> List[str]:
    categories: List[str] = []
    for tag in area_tags:
        normalized = _normalize_space(tag)
        lowered = normalized.lower()
        if "gen ed" in lowered or "gen-ed" in lowered or "gened" in lowered:
            category = re.sub(r"\bgen[\s-]?ed\b", "", normalized, flags=re.IGNORECASE)
            cleaned = _normalize_space(category)
            categories.append(cleaned if cleaned else normalized)
            continue
        if lowered in _GEN_ED_STANDALONE_TAGS:
            categories.append(normalized)
    return list(dict.fromkeys(categories))


def _is_wic(area_tags: List[str], notes_value: object) -> bool:
    for tag in area_tags:
        lowered = tag.lower()
        if "writing intensive course" in lowered or lowered == "wic":
            return True
    notes = _raw_cell(notes_value).lower()
    return "writing intensive course" in notes


def _row_value(row: List[object], idx: Optional[int]) -> object:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return row[idx]


def _collect_availability_fields(
    row: List[object],
    availability_indices: Dict[str, int],
) -> Dict[str, List[str]]:
    fields: Dict[str, List[str]] = {}
    for header, idx in availability_indices.items():
        tags = _split_tags(_row_value(row, idx))
        if tags:
            fields[header] = tags
    return fields


def _iter_xlsx_rows(path: Path) -> tuple[List[str], List[List[object]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - dependency error
        raise RuntimeError(f"openpyxl is required to read {path.suffix} files: {exc}") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers = [_clean_header(h) for h in rows[0]]
        body = [list(r) for r in rows[1:]]
        return headers, body
    finally:
        wb.close()


def _iter_xls_rows(path: Path) -> tuple[List[str], List[List[object]]]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:  # pragma: no cover - dependency error
        raise RuntimeError(f"xlrd is required to read {path.suffix} files: {exc}") from exc

    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        return [], []
    headers = [_clean_header(h) for h in sheet.row_values(0)]
    body = [sheet.row_values(r) for r in range(1, sheet.nrows)]
    return headers, body


def _load_raw_rows(path: Path) -> tuple[List[str], List[List[object]]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _iter_xlsx_rows(path)
    if suffix == ".xls":
        return _iter_xls_rows(path)
    raise RuntimeError(f"Unsupported Excel file format: {path.suffix}")


def _header_lookup(headers: List[str]) -> Dict[str, int]:
    normalized = [_normalize_header_key(h) for h in headers]
    field_to_idx: Dict[str, int] = {}
    for field, aliases in _HEADER_ALIASES.items():
        for idx, key in enumerate(normalized):
            if key in aliases:
                field_to_idx[field] = idx
                break
    return field_to_idx


def _availability_indices(headers: List[str], reserved: set[int]) -> Dict[str, int]:
    indices: Dict[str, int] = {}
    for idx, header in enumerate(headers):
        if idx in reserved:
            continue
        normalized = _normalize_header_key(header)
        if any(hint in normalized for hint in _TERM_HEADER_HINTS):
            indices[_clean_header(header)] = idx
    return indices


def _clone_course(course: Dict[str, Any]) -> Dict[str, Any]:
    clone = dict(course)
    clone["area_of_study_tags"] = list(course.get("area_of_study_tags") or [])
    clone["tags"] = list(course.get("tags") or [])
    clone["gen_ed_tags"] = list(course.get("gen_ed_tags") or [])
    clone["semester_availability"] = list(course.get("semester_availability") or [])
    availability_fields = course.get("availability_fields") or {}
    clone["availability_fields"] = {
        key: list(values)
        for key, values in availability_fields.items()
        if isinstance(values, list)
    }
    clone.pop("_search_blob", None)
    return clone


def _merge_course(existing: Dict[str, Any], incoming: Dict[str, Any]) -> Dict[str, Any]:
    merged = dict(existing)
    if (not merged.get("title") or merged.get("title") == merged.get("code")) and incoming.get("title"):
        merged["title"] = incoming["title"]
    if merged.get("credits") is None and incoming.get("credits") is not None:
        merged["credits"] = incoming["credits"]
    if not merged.get("level") and incoming.get("level"):
        merged["level"] = incoming["level"]
    merged["area_of_study_tags"] = list(
        dict.fromkeys((merged.get("area_of_study_tags") or []) + (incoming.get("area_of_study_tags") or []))
    )
    merged["tags"] = list(dict.fromkeys((merged.get("tags") or []) + (incoming.get("tags") or [])))
    merged["gen_ed_tags"] = list(dict.fromkeys((merged.get("gen_ed_tags") or []) + (incoming.get("gen_ed_tags") or [])))
    merged["wic"] = bool(merged.get("wic") or incoming.get("wic"))
    merged["semester_availability"] = list(
        dict.fromkeys((merged.get("semester_availability") or []) + (incoming.get("semester_availability") or []))
    )

    fields: Dict[str, List[str]] = {}
    for source in (merged.get("availability_fields") or {}, incoming.get("availability_fields") or {}):
        for key, values in source.items():
            if not isinstance(values, list):
                continue
            fields[key] = list(dict.fromkeys((fields.get(key) or []) + values))
    merged["availability_fields"] = fields
    merged["_search_blob"] = _build_search_blob(merged)
    return merged


def _build_search_blob(course: Dict[str, Any]) -> str:
    parts: List[str] = [
        course.get("code", ""),
        course.get("title", ""),
        " ".join(course.get("tags") or []),
        " ".join(course.get("area_of_study_tags") or []),
        " ".join(course.get("gen_ed_tags") or []),
        " ".join(course.get("semester_availability") or []),
    ]
    return _normalize_space(" ".join([str(p) for p in parts if p])).lower()


def _normalize_course_from_payload(raw_course: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not isinstance(raw_course, dict):
        return None

    code = normalize_course_code(str(raw_course.get("code") or ""))
    if not code:
        return None

    title = _clean_cell(raw_course.get("title")) or code
    department = _clean_cell(raw_course.get("department")) or code.split(" ", 1)[0]
    prefix = _clean_cell(raw_course.get("prefix")) or department
    level = _clean_cell(raw_course.get("level"))

    credits_raw = raw_course.get("credits")
    credits = None
    if isinstance(credits_raw, bool):
        credits = None
    elif isinstance(credits_raw, (int, float)) and credits_raw > 0:
        credits = int(round(float(credits_raw)))
    else:
        parsed_credits = _parse_numeric(credits_raw)
        if parsed_credits is not None and parsed_credits > 0:
            credits = int(round(parsed_credits))

    area_of_study_tags = [
        _normalize_space(str(tag))
        for tag in (raw_course.get("area_of_study_tags") or [])
        if _normalize_space(str(tag))
    ]
    tags = [
        _normalize_space(str(tag))
        for tag in (raw_course.get("tags") or area_of_study_tags)
        if _normalize_space(str(tag))
    ]
    gen_ed_tags = [
        _normalize_space(str(tag))
        for tag in (raw_course.get("gen_ed_tags") or [])
        if _normalize_space(str(tag))
    ]
    semester_availability = [
        _normalize_space(str(term))
        for term in (raw_course.get("semester_availability") or [])
        if _normalize_space(str(term))
    ]

    availability_fields_raw = raw_course.get("availability_fields") or {}
    availability_fields: Dict[str, List[str]] = {}
    if isinstance(availability_fields_raw, dict):
        for key, values in availability_fields_raw.items():
            cleaned_key = _clean_cell(key)
            if not cleaned_key or not isinstance(values, list):
                continue
            cleaned_values = [
                _normalize_space(str(value))
                for value in values
                if _normalize_space(str(value))
            ]
            if cleaned_values:
                availability_fields[cleaned_key] = list(dict.fromkeys(cleaned_values))

    record = {
        "code": code,
        "title": title,
        "credits": credits,
        "department": department,
        "prefix": prefix,
        "level": level,
        "area_of_study_tags": list(dict.fromkeys(area_of_study_tags)),
        "tags": list(dict.fromkeys(tags or area_of_study_tags)),
        "gen_ed_tags": list(dict.fromkeys(gen_ed_tags)),
        "wic": bool(raw_course.get("wic")),
        "semester_availability": list(dict.fromkeys(semester_availability)),
        "availability_fields": availability_fields,
    }
    record["_search_blob"] = _normalize_space(str(raw_course.get("_search_blob") or "")) or _build_search_blob(record)
    return record


def _normalize_catalog_payload(data: Dict[str, Any]) -> Dict[str, Any]:
    by_code_raw = data.get("by_code")
    course_entries: List[Dict[str, Any]] = []
    if isinstance(by_code_raw, dict):
        for raw_course in by_code_raw.values():
            if isinstance(raw_course, dict):
                course_entries.append(raw_course)
    else:
        for raw_course in data.get("courses") or []:
            if isinstance(raw_course, dict):
                course_entries.append(raw_course)

    normalized_by_code: Dict[str, Dict[str, Any]] = {}
    for raw_course in course_entries:
        course = _normalize_course_from_payload(raw_course)
        if course is None:
            continue
        existing = normalized_by_code.get(course["code"])
        normalized_by_code[course["code"]] = _merge_course(existing, course) if existing else course

    courses_sorted = [normalized_by_code[code] for code in sorted(normalized_by_code.keys())]
    data["by_code"] = normalized_by_code
    data["courses"] = courses_sorted
    data["codes"] = [course["code"] for course in courses_sorted]
    return data


def _apply_catalog_payload(
    payload: Dict[str, Any],
    *,
    source_label: Optional[str],
    source_mtime: Optional[float],
) -> int:
    global courses_by_code, _courses_sorted, _loaded_path, _loaded_mtime, _catalog_payload

    normalized = _normalize_catalog_payload(payload)
    courses_by_code = normalized["by_code"]
    _courses_sorted = normalized["courses"]
    _loaded_path = source_label
    _loaded_mtime = source_mtime
    _catalog_payload = normalized
    return len(courses_by_code)


def _build_course_record(
    row: List[object],
    headers: List[str],
    lookup: Dict[str, int],
    availability_indices: Dict[str, int],
) -> Optional[Dict[str, Any]]:
    department_value = _row_value(row, lookup.get("department"))
    course_value = _row_value(row, lookup.get("course"))
    code_value = _row_value(row, lookup.get("code"))
    code = _code_from_parts(_clean_cell(department_value), _clean_cell(course_value), _clean_cell(code_value))
    if not code:
        return None

    title = _clean_cell(_row_value(row, lookup.get("title"))) or code
    level = _clean_cell(_row_value(row, lookup.get("level")))
    area_tags = _split_tags(_row_value(row, lookup.get("area_of_study")))
    notes_value = _row_value(row, lookup.get("course_notes"))
    credits = _credits_from_value_or_notes(_row_value(row, lookup.get("credits")), notes_value)

    availability_fields = _collect_availability_fields(row, availability_indices)
    availability_flat: List[str] = []
    for values in availability_fields.values():
        availability_flat.extend(values)
    semester_availability = list(dict.fromkeys(availability_flat))

    prefix = code.split(" ", 1)[0]
    record = {
        "code": code,
        "title": title,
        "credits": credits,
        "department": prefix,
        "prefix": prefix,
        "level": level,
        "area_of_study_tags": area_tags,
        "tags": list(area_tags),
        "gen_ed_tags": _gen_ed_from_area_tags(area_tags),
        "wic": _is_wic(area_tags, notes_value),
        "semester_availability": semester_availability,
        "availability_fields": availability_fields,
    }
    record["_search_blob"] = _build_search_blob(record)
    return record


def build_course_catalog_data(path: Path | str) -> Dict[str, Any]:
    source = Path(path)
    if not source.exists():
        raise RuntimeError(f"Excel course catalog not found: {source}")

    mtime = source.stat().st_mtime
    headers, rows = _load_raw_rows(source)
    if not headers:
        raise RuntimeError(f"Excel course catalog is empty: {source}")

    lookup = _header_lookup(headers)
    if "title" not in lookup:
        raise RuntimeError("Excel catalog is missing a title/label column.")
    if "department" not in lookup and "code" not in lookup:
        raise RuntimeError("Excel catalog is missing department/code columns.")
    if "course" not in lookup and "code" not in lookup:
        raise RuntimeError("Excel catalog is missing course number/code columns.")

    reserved = set(lookup.values())
    availability_indices = _availability_indices(headers, reserved)

    parsed: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        record = _build_course_record(row, headers, lookup, availability_indices)
        if record is None:
            continue
        code = record["code"]
        existing = parsed.get(code)
        parsed[code] = _merge_course(existing, record) if existing else record

    return {
        "courses": [parsed[code] for code in sorted(parsed.keys())],
        "by_code": parsed,
        "codes": list(sorted(parsed.keys())),
        "source_path": str(source),
        "source_mtime": mtime,
    }


def load_course_catalog(path: Path | str) -> int:
    source = Path(path)
    mtime = source.stat().st_mtime
    if _loaded_path == str(source) and _loaded_mtime == mtime and courses_by_code:
        return len(courses_by_code)

    payload = build_course_catalog_data(source)
    count = _apply_catalog_payload(
        payload,
        source_label=str(source),
        source_mtime=mtime,
    )
    logging.info("Loaded Excel course catalog from %s (%s courses).", source, len(courses_by_code))
    return count


def load_course_catalog_from_data(data: Dict[str, Any], source_label: str = "excel_catalog.json") -> int:
    return _apply_catalog_payload(
        data,
        source_label=source_label,
        source_mtime=None,
    )


def get_catalog_data() -> Dict[str, Any]:
    return _catalog_payload


def _term_matches(course: Dict[str, Any], term: Optional[str]) -> bool:
    if not term:
        return True
    needle = _normalize_space(term).lower()
    if not needle:
        return True
    for value in course.get("semester_availability") or []:
        if needle in _normalize_space(str(value)).lower():
            return True
    for values in (course.get("availability_fields") or {}).values():
        for value in values:
            if needle in _normalize_space(str(value)).lower():
                return True
    return False


def get_course(code: str) -> Optional[Dict[str, Any]]:
    normalized = normalize_course_code(code)
    course = courses_by_code.get(normalized)
    if not course:
        return None
    return _clone_course(course)


def get_course_codes() -> set[str]:
    return set(courses_by_code.keys())


def list_courses(term: Optional[str] = None) -> List[Dict[str, Any]]:
    return [_clone_course(course) for course in _courses_sorted if _term_matches(course, term)]


def _search_rank(course: Dict[str, Any], query: str) -> Optional[int]:
    query_clean = _normalize_space(query).lower()
    query_flat = query_clean.replace(" ", "")
    if not query_clean:
        return 0

    code = str(course.get("code", "")).lower()
    code_flat = code.replace(" ", "")
    title = str(course.get("title", "")).lower()
    blob = str(course.get("_search_blob", "")).lower()

    if code_flat == query_flat:
        return 0
    if code_flat.startswith(query_flat):
        return 1
    if query_flat in code_flat:
        return 2
    if title.startswith(query_clean):
        return 3
    if query_clean in title:
        return 4
    if query_clean in blob:
        return 5
    return None


def search_courses(query: str, term: Optional[str] = None, limit: int = 50) -> List[Dict[str, Any]]:
    if not query or not str(query).strip():
        return []

    bounded_limit = max(1, min(int(limit), 200))
    ranked: List[tuple[int, str, Dict[str, Any]]] = []
    for course in _courses_sorted:
        if not _term_matches(course, term):
            continue
        rank = _search_rank(course, query)
        if rank is None:
            continue
        ranked.append((rank, course.get("code", ""), course))

    ranked.sort(key=lambda item: (item[0], item[1]))
    return [_clone_course(item[2]) for item in ranked[:bounded_limit]]
