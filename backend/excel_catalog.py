from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Set
import logging
import re

PROGRAM_TAG_ALIASES: Dict[str, List[str]] = {
    "Business Administration": ["BUS"],
    "Computer Science": ["COS", "CS"],
    "Economics": ["ECO"],
    "European Studies": ["EUR"],
    "Finance": ["FIN"],
    "History and Civilizations": ["HC", "HTY"],
    "Information Systems": ["IS", "ISM"],
    "Journalism and Mass Communication": ["JMC"],
    "Literature": ["LIT", "ENG"],
    "Mathematics": ["MAT"],
    "Modern Languages and Cultures": ["MLC"],
    "Physics": ["PHY"],
    "Political Science and International Relations": ["POS"],
    "Psychology": ["PSY"],
    "Film and Creative Media": ["Film", "FIL"],
    "Sustainability Studies": ["Sustainability", "Sustainabiliy"],
}

TARGET_CASE_STUDIES_TAG = "Case Studies in Textual Analysis Gen Ed"
TARGET_WIC_TAG = "Writing Intensive Course"

_EXCEL_CACHE: Dict[str, object] = {
    "path": None,
    "mtime": None,
    "data": None,
}


def _empty_catalog() -> Dict[str, object]:
    return {
        "courses": [],
        "by_code": {},
        "codes": set(),
    }


def _clean_header(value: object) -> str:
    text = str(value or "")
    return re.sub(r"\s+", " ", text).strip()


def _clean_cell(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value or "")
    return re.sub(r"\s+", " ", text).strip()


def _split_tags(value: object) -> List[str]:
    text = str(value or "")
    if not text.strip():
        return []
    lines = re.split(r"[\r\n]+", text)
    tags: List[str] = []
    for line in lines:
        cleaned = re.sub(r"\s+", " ", line).strip()
        if cleaned:
            tags.append(cleaned)
    return list(dict.fromkeys(tags))


def _load_xlsx(path: Path) -> Dict[str, object]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        logging.warning("openpyxl is unavailable; cannot read Excel catalog: %s", exc)
        return _empty_catalog()

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        logging.warning("Failed to open Excel catalog %s: %s", path, exc)
        return _empty_catalog()

    try:
        sheet = wb.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            return _empty_catalog()

        header_list = [_clean_header(h) for h in headers]
        header_map = {h: i for i, h in enumerate(header_list)}
        dept_idx = header_map.get("Department")
        course_idx = header_map.get("Course")
        area_idx = header_map.get("Area of Study")
        if dept_idx is None or course_idx is None or area_idx is None:
            return _empty_catalog()

        entries: Dict[str, Dict[str, object]] = {}
        for row in rows:
            dept = _clean_cell(row[dept_idx]) if dept_idx < len(row) else ""
            course = _clean_cell(row[course_idx]) if course_idx < len(row) else ""
            area = row[area_idx] if area_idx < len(row) else ""
            if not dept or not course:
                continue
            department = dept.upper()
            number = course
            code = f"{department} {number}".strip().upper()
            tags = _split_tags(area)
            existing = entries.get(code)
            if existing is None:
                entries[code] = {
                    "code": code,
                    "department": department,
                    "number": number,
                    "tags": tags,
                }
            else:
                merged = list(dict.fromkeys((existing.get("tags") or []) + tags))
                existing["tags"] = merged

        courses = sorted(entries.values(), key=lambda c: c["code"])
        return {
            "courses": courses,
            "by_code": entries,
            "codes": set(entries.keys()),
        }
    finally:
        wb.close()


def _load_xls(path: Path) -> Dict[str, object]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        logging.warning("xlrd is unavailable; cannot read legacy Excel catalog: %s", exc)
        return _empty_catalog()

    try:
        book = xlrd.open_workbook(str(path))
    except Exception as exc:
        logging.warning("Failed to open legacy Excel catalog %s: %s", path, exc)
        return _empty_catalog()

    sheet = book.sheet_by_index(0)
    if sheet.nrows < 1:
        return _empty_catalog()

    headers = [_clean_header(h) for h in sheet.row_values(0)]
    header_map = {h: i for i, h in enumerate(headers)}
    dept_idx = header_map.get("Department")
    course_idx = header_map.get("Course")
    area_idx = header_map.get("Area of Study")
    if dept_idx is None or course_idx is None or area_idx is None:
        return _empty_catalog()

    entries: Dict[str, Dict[str, object]] = {}
    for r in range(1, sheet.nrows):
        row = sheet.row_values(r)
        dept = _clean_cell(row[dept_idx]) if dept_idx < len(row) else ""
        course = _clean_cell(row[course_idx]) if course_idx < len(row) else ""
        area = row[area_idx] if area_idx < len(row) else ""
        if not dept or not course:
            continue
        department = dept.upper()
        number = course
        code = f"{department} {number}".strip().upper()
        tags = _split_tags(area)
        existing = entries.get(code)
        if existing is None:
            entries[code] = {
                "code": code,
                "department": department,
                "number": number,
                "tags": tags,
            }
        else:
            merged = list(dict.fromkeys((existing.get("tags") or []) + tags))
            existing["tags"] = merged

    courses = sorted(entries.values(), key=lambda c: c["code"])
    return {
        "courses": courses,
        "by_code": entries,
        "codes": set(entries.keys()),
    }


def load_excel_catalog(path: Path | None) -> Dict[str, object]:
    if path is None or not path.exists():
        return _empty_catalog()

    cache_path = _EXCEL_CACHE.get("path")
    cache_mtime = _EXCEL_CACHE.get("mtime")
    mtime = path.stat().st_mtime
    if cache_path == str(path) and cache_mtime == mtime and _EXCEL_CACHE.get("data") is not None:
        return _EXCEL_CACHE["data"]  # type: ignore[return-value]

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        data = _load_xlsx(path)
    elif suffix == ".xls":
        data = _load_xls(path)
    else:
        logging.warning("Unsupported Excel catalog format: %s", path)
        data = _empty_catalog()

    _EXCEL_CACHE["path"] = str(path)
    _EXCEL_CACHE["mtime"] = mtime
    _EXCEL_CACHE["data"] = data
    return data


def get_excel_course_codes(excel_catalog: Dict[str, object]) -> Set[str]:
    codes = excel_catalog.get("codes")
    if isinstance(codes, set):
        return set(codes)
    return set(excel_catalog.get("by_code", {}).keys())


def _program_tag_prefixes(programs: List[str]) -> Set[str]:
    prefixes: Set[str] = set()
    for name in programs:
        aliases = PROGRAM_TAG_ALIASES.get(name)
        if aliases:
            prefixes.update(aliases)
    return prefixes


def _tag_starts_with_prefix(tag: str, prefix: str) -> bool:
    t = re.sub(r"\s+", " ", tag).strip().lower()
    p = re.sub(r"\s+", " ", prefix).strip().lower()
    return t.startswith(p + " ")


def _tag_has_words(tag: str, *needles: str) -> bool:
    t = re.sub(r"\s+", " ", tag).strip().lower()
    return all(n.lower() in t for n in needles)


def get_recommended_electives(
    excel_catalog: Dict[str, object],
    selected_majors: List[str],
    selected_minors: List[str],
) -> List[Dict[str, object]]:
    major_prefixes = _program_tag_prefixes(selected_majors)
    minor_prefixes = _program_tag_prefixes(selected_minors)
    if not major_prefixes and not minor_prefixes:
        return []

    results: List[Dict[str, object]] = []
    by_code: Dict[str, Dict[str, object]] = excel_catalog.get("by_code", {})  # type: ignore[assignment]
    for code in sorted(by_code.keys()):
        entry = by_code[code]
        tags = entry.get("tags") or []
        if not isinstance(tags, list) or not tags:
            continue

        matched_major: List[str] = []
        matched_minor: List[str] = []
        exclude = False

        for tag in tags:
            if not isinstance(tag, str):
                continue
            for prefix in major_prefixes:
                if _tag_starts_with_prefix(tag, prefix) and _tag_has_words(tag, "major", "required"):
                    exclude = True
                    break
            if exclude:
                break
            for prefix in major_prefixes:
                if (
                    _tag_starts_with_prefix(tag, prefix)
                    and _tag_has_words(tag, "major", "elective")
                    and not _tag_has_words(tag, "required")
                ):
                    matched_major.append(tag)
            for prefix in minor_prefixes:
                if _tag_starts_with_prefix(tag, prefix) and _tag_has_words(tag, "minor", "elective"):
                    matched_minor.append(tag)

        if exclude:
            continue

        if not matched_major and not matched_minor:
            continue

        matched_major = list(dict.fromkeys(matched_major))
        matched_minor = list(dict.fromkeys(matched_minor))

        results.append({
            "code": code,
            "tags": list(dict.fromkeys([t for t in tags if isinstance(t, str)])),
            "matched_major_tags": matched_major,
            "matched_minor_tags": matched_minor,
        })

    return results


def get_case_studies_gened_courses(excel_catalog: Dict[str, object]) -> List[str]:
    by_code: Dict[str, Dict[str, object]] = excel_catalog.get("by_code", {})  # type: ignore[assignment]
    target = re.sub(r"\s+", " ", TARGET_CASE_STUDIES_TAG).strip().lower()
    matches: List[str] = []
    for code in sorted(by_code.keys()):
        tags = by_code[code].get("tags") or []
        if not isinstance(tags, list):
            continue
        for tag in tags:
            if not isinstance(tag, str):
                continue
            normalized = re.sub(r"\s+", " ", tag).strip().lower()
            if normalized == target:
                matches.append(code)
                break
    return matches


def get_selected_program_elective_tags(
    excel_catalog: Dict[str, object],
    selected_majors: List[str],
    selected_minors: List[str],
) -> Dict[str, List[str]]:
    major_prefixes = _program_tag_prefixes(selected_majors)
    minor_prefixes = _program_tag_prefixes(selected_minors)
    selected_prefixes = major_prefixes | minor_prefixes
    if not selected_prefixes:
        return {}

    by_code: Dict[str, Dict[str, object]] = excel_catalog.get("by_code", {})  # type: ignore[assignment]
    output: Dict[str, List[str]] = {}
    for code, entry in by_code.items():
        tags = entry.get("tags") or []
        if not isinstance(tags, list) or not tags:
            continue
        matched: List[str] = []
        for tag in tags:
            if not isinstance(tag, str):
                continue
            if not any(_tag_starts_with_prefix(tag, prefix) for prefix in selected_prefixes):
                continue
            if _tag_has_words(tag, "major", "elective") or _tag_has_words(tag, "minor", "elective"):
                matched.append(tag)
        if matched:
            output[code] = list(dict.fromkeys(matched))
    return output


def compute_catalog_integrity(pdf_codes: Set[str], excel_codes: Set[str]) -> Dict[str, List[str]]:
    excel_only = sorted(excel_codes - pdf_codes)
    pdf_only = sorted(pdf_codes - excel_codes)
    return {
        "excel_only": excel_only,
        "pdf_only": pdf_only,
    }
